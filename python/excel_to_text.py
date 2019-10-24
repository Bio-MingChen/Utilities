#!/usr/bin/env python
# -*- coding=utf-8 -*-

#Extracting indicated sheet to plain text
import sys
if sys.version.startswith('2.'):
    reload(sys)
    sys.setdefaultencoding('utf-8')
import argparse
import datetime
from openpyxl import load_workbook


time = datetime.datetime.now().strftime('%Y-%m-%d')

def excel_to_txt(args):
    excel = args.get('excel')
    sheet_name = args.get('sheet_name') if args.get('sheet_name') else ''
    sheet_number = args.get('sheet_number') if args.get('sheet_number') else ''

    book = load_workbook(excel)
    sheet_list = []
    if sheet_number:
        sheet_list += [book.sheetnames[int(i)] for i in sheet_number.strip().split(',')]
    if sheet_name and (sheet_name not in sheet_list):
        sheet_list += [sheet_name]
    if sheet_list:
        for sheet in sheet_list:
            ofile = sheet + time + '.xls'
            print('Writing %s to %s' % (sheet,ofile))
            with open(ofile,'w') as odata:
                for row in book[sheet]:
                    row_list = [cell.value for cell in row]
                    row_list = ['' if i is None else i for i in row_list]
                    odata.write("\t".join(row_list) + "\n")
        print('Writing Completion.') 
    else:
        raise IndexError('sheet number out of range!')

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
        Example:
        
            excel_to_txt excel  # extract all sheets 
            excel_to_txt excel --sheet_number 0,1,2
            excel_to_txt excel --sheet_name 'xxx'
        '''
    )
    parser.add_argument('excel',help='Excel Name')
    parser.add_argument('--sheet_number','-num',help='The index of sheet number zero-based')
    parser.add_argument('--sheet_name','-name',help='The sheet name')
    args = parser.parse_args()
    excel_to_txt(vars(args))
